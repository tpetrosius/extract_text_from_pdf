import openpyxl, datetime
from openpyxl import Workbook


def create_final_report():
	"""Creates excel file to store all extracted data."""
	
	file_name = 'Final Report ' + str(datetime.date.today()) + '.xlsx'
	folder_path = r'C:\Users\Vartotojas\Desktop'
	file_path = folder_path + '\\' + file_name
	
	column_names = [('A1','Client Name'), ('B1','Nationality'), ('C1','Date of birth'), ('D1','Client number'), 
					('E1','Region code'), ('F1','Number of products'), ('G1','Transaction volume YTD')]
	
	wb = Workbook()
	
	column_name_number = 0
	
	#Create Corporate clients sheet and write columns
	sheet_1 = wb.create_sheet('Corporate Clients', 0)

	for i in column_names:
		column_cell = column_names[column_name_number][0]
		column_name = column_names[column_name_number][1]
		sheet_1[column_cell] = column_name
		column_name_number += 1
	
	column_name_number = 0
	
	#Create Retail clients sheet and write columns
	sheet_2 = wb.create_sheet('Retail Clients', 1)
	
	for i in column_names:
		column_cell = column_names[column_name_number][0]
		column_name = column_names[column_name_number][1]
		sheet_2[column_cell] = column_name
		column_name_number += 1
	
	column_name_number = 0
	
	#Create Exceptions sheet and write columns
	sheet_3 = wb.create_sheet('Exceptions', 2)
	for i in column_names:
		column_cell = column_names[column_name_number][0]
		column_name = column_names[column_name_number][1]
		sheet_3[column_cell] = column_name
		column_name_number += 1
	
	column_name_number = 0
	
	if 'Sheet' in wb.sheetnames:
		del wb['Sheet']
	
	wb.save(file_path)
	
	return file_path
