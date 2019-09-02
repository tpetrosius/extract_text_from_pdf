import openpyxl

def write_data(final_report, values, SEK_rate):
	"""Writes extracted data to final report"""
	
	wb = openpyxl.load_workbook(final_report)
	
	# Check if there is any missing value, if there is it is exception
	missing = 0
		
	for value in values.values():
		if value == "":
			missing += 1
		else:
			missing += 0
		
	row_number = 0
		
	if missing == 0:
		if values["Corporate account"] == "Yes":
			sheet = wb["Corporate Clients"]
			row_number = sheet.max_row + 1
		else:
			sheet = wb["Retail Clients"]
			row_number = sheet.max_row + 1
	else:
		sheet = wb["Exceptions"]
		row_number = sheet.max_row + 1
		
	#Write values to final report (excel template).
	# Name.
	sheet["A" + str(row_number)] = values["Name"]
		
	# Nationality.
	sheet["B" + str(row_number)] = values["Nationality"]
			
	# Date of birth.
	sheet["C" + str(row_number)] = values["Date of birth"]
		
	# Customer number.
	sheet["D" + str(row_number)] = values["Client number"]
		
	# Region code
	sheet["E" + str(row_number)] = values["Region code"]
		
	# Number of products
	sheet["F" + str(row_number)] = values["Number of products"]
		
	# Calculate EUR and Transaction volume YTD
	volume = values["Transaction volume YTD"]
	if volume.find("EK") > 0:
		amount = volume.replace("SEK ", "")
		amount_2 = amount.replace(".", "")
		amount_3 = amount_2.replace(",", ".")
		EUR = float(amount_3) / float(SEK_rate)
		sheet["G" + str(row_number)] = str(round(EUR, 2)) + " EUR"
		
	wb.save(final_report)
